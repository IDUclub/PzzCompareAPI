"""Read a tabular zone-description file (CSV/XLSX) into rows + header names.

Users hand ПЗЗ zone descriptions as spreadsheets, not JSON. This adapter turns
either format into ``(rows, headers)`` where each row is a ``{header: str}`` dict,
so the same column-detection and conversion logic serves both.

Encoding is the sharp edge: Russian CSVs exported from Excel are commonly
Windows-1251 with a ``;`` delimiter, not UTF-8/comma. Decoding blind mangles the
Cyrillic, so the CSV path sniffs both encoding and delimiter. XLSX is XML/UTF-8
internally — ``openpyxl`` yields proper ``str`` with no encoding guesswork.
"""
from __future__ import annotations

import csv
import io
from typing import Any


class TableReadError(ValueError):
    """The uploaded file could not be read as a table (bad format/encoding/empty)."""


_CSV_ENCODINGS = ("utf-8-sig", "cp1251")
_CSV_DELIMITERS = ";,\t"


def read_table(
    data: bytes, filename: str, *, sheet: str | None = None
) -> tuple[list[dict[str, str]], list[str]]:
    """Read ``data`` as a table, dispatching on ``filename`` extension.

    Returns ``(rows, headers)``. The first non-empty row is the header; every
    subsequent row becomes a ``{header: value}`` dict (missing cells -> ""). Raises
    :class:`TableReadError` on an unsupported extension, an undecodable CSV, an
    empty sheet, or a missing named ``sheet``.
    """
    name = (filename or "").lower()
    if name.endswith(".xlsx"):
        return _read_xlsx(data, sheet=sheet)
    if name.endswith(".csv"):
        return _read_csv(data)
    raise TableReadError(
        "unsupported table format: expected .csv or .xlsx, got "
        f"«{filename}»"
    )


def _decode_csv(data: bytes) -> str:
    for enc in _CSV_ENCODINGS:
        try:
            return data.decode(enc)
        except UnicodeDecodeError:
            continue
    raise TableReadError(
        "cannot decode CSV: tried " + ", ".join(_CSV_ENCODINGS)
    )


def _sniff_delimiter(sample: str) -> str:
    try:
        return csv.Sniffer().sniff(sample, delimiters=_CSV_DELIMITERS).delimiter
    except csv.Error:
        # Russian Excel exports default to ';'; fall back to it when the sniffer
        # can't decide (e.g. a single-column file).
        return ";" if ";" in sample else ","


def _read_csv(data: bytes) -> tuple[list[dict[str, str]], list[str]]:
    text = _decode_csv(data)
    if not text.strip():
        raise TableReadError("CSV file is empty")
    delimiter = _sniff_delimiter(text[:4096])
    reader = csv.reader(io.StringIO(text), delimiter=delimiter)
    matrix = [row for row in reader if any((c or "").strip() for c in row)]
    return _matrix_to_rows(matrix)


def _read_xlsx(
    data: bytes, *, sheet: str | None
) -> tuple[list[dict[str, str]], list[str]]:
    from openpyxl import load_workbook

    try:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl raises assorted types on bad input
        raise TableReadError(f"cannot open XLSX: {exc}") from exc
    try:
        if sheet is not None:
            if sheet not in wb.sheetnames:
                raise TableReadError(
                    f"sheet «{sheet}» not found; available: {', '.join(wb.sheetnames)}"
                )
            ws = wb[sheet]
        else:
            ws = wb.active
        matrix: list[list[str]] = []
        for raw_row in ws.iter_rows(values_only=True):
            cells = ["" if v is None else str(v).strip() for v in raw_row]
            if any(cells):
                matrix.append(cells)
    finally:
        wb.close()
    return _matrix_to_rows(matrix)


def _matrix_to_rows(
    matrix: list[list[Any]],
) -> tuple[list[dict[str, str]], list[str]]:
    if not matrix:
        raise TableReadError("table has no data rows")
    raw_headers = [str(h).strip() for h in matrix[0]]
    headers = _dedupe_headers(raw_headers)
    rows: list[dict[str, str]] = []
    for cells in matrix[1:]:
        row = {
            headers[i]: (str(cells[i]).strip() if i < len(cells) and cells[i] is not None else "")
            for i in range(len(headers))
        }
        rows.append(row)
    return rows, headers


def _dedupe_headers(headers: list[str]) -> list[str]:
    """Ensure header names are unique and non-empty so they can key row dicts."""
    seen: dict[str, int] = {}
    result: list[str] = []
    for i, h in enumerate(headers):
        name = h or f"column_{i + 1}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        result.append(name)
    return result
